import os
import shutil
import sys
from datetime import datetime
from pathlib import Path

# Add project root to sys.path to resolve 'app' module
project_root = str(Path(__file__).resolve().parent.parent)
if project_root not in sys.path:
    sys.path.append(project_root)


from openpyxl import Workbook, load_workbook  # noqa: E402
from openpyxl.styles import Font  # noqa: E402

# Import centralized schema
from app.services.logger_service import (EVAL_COLUMNS,  # noqa: E402
                                         EVAL_SHEET, PROD_COLUMNS, PROD_SHEET)

log_dir = "logs"
log_path = f"{log_dir}/response_log.xlsx"
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
os.makedirs(log_dir, exist_ok=True)

# --- Archive the existing file ---
if os.path.exists(log_path):
    counts = {}
    try:
        wb_old = load_workbook(log_path)
        for sheet in wb_old.sheetnames:
            ws = wb_old[sheet]
            counts[sheet] = ws.max_row - 1
    except Exception:
        counts = {"(unreadable)": "?"}

    archive_path = f"{log_dir}/response_log_{timestamp}.xlsx"
    shutil.move(log_path, archive_path)
    summary = ", ".join([f"{s}: {n} entries" for s, n in counts.items()])
    print(f"  Archived -> {archive_path}")
    print(f"  ({summary})")
else:
    print("  No existing log found. Creating fresh one.")

# --- Create a fresh workbook using Master Schema ---
wb = Workbook()

# Sheet 1: Production
ws_prod = wb.active
ws_prod.title = PROD_SHEET
ws_prod.append(PROD_COLUMNS)
for cell in ws_prod[1]:
    cell.font = Font(bold=True)

# Sheet 2: Evaluation
ws_eval = wb.create_sheet(EVAL_SHEET)
ws_eval.append(EVAL_COLUMNS)
for cell in ws_eval[1]:
    cell.font = Font(bold=True)

wb.save(log_path)
print(f"\n✓ Fresh log created: {log_path}")
print("  Integrated with centralized Master Schema definition.")
print("  Ready for new queries and evaluations!")
