import os
import threading
import shutil
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# --- Centralized Schema Source of Truth ---
PROD_SHEET     = "Production"
PROD_COLUMNS   = [
    "Timestamp", "User Query", "Bot Response", "Time Taken (s)",
    "Session ID", "Source", "Context"
]

EVAL_SHEET     = "Evaluation"
EVAL_COLUMNS   = [
    "Timestamp", "Prompt", "Bot Answer", "Source",
    "Retrieval Confidence (%)", "Latency (s)", "Server Time (s)",
    "BERTScore F1 %", "Link Validity", "Accuracy %", "Context"
]

class ResponseLogger:
    def __init__(self, log_dir=None, filename="response_log.xlsx"):
        """Initialize with thread safety and self-healing multi-sheet file."""
        self.lock = threading.Lock()
        self.log_dir  = Path(log_dir) if log_dir else Path(__file__).resolve().parent.parent.parent / "logs"
        self.log_file = self.log_dir / filename
        self._ensure_log_file()

    def _make_fresh_wb(self):
        """Create a brand new workbook with synchronized sheets."""
        wb = Workbook()
        # Production sheet
        ws_prod = wb.active
        ws_prod.title = PROD_SHEET
        ws_prod.append(PROD_COLUMNS)
        for cell in ws_prod[1]:
            cell.font = Font(bold=True)

        # Evaluation sheet
        ws_eval = wb.create_sheet(EVAL_SHEET)
        ws_eval.append(EVAL_COLUMNS)
        for cell in ws_eval[1]:
            cell.font = Font(bold=True)
        return wb

    def _ensure_log_file(self):
        """Self-healing header check: Archives the log if columns are outdated."""
        with self.lock:
            self.log_dir.mkdir(parents=True, exist_ok=True)
            
            needs_reset = False
            if not self.log_file.exists():
                needs_reset = True
            else:
                try:
                    wb = load_workbook(self.log_file)
                    # Check PROD sheet
                    if PROD_SHEET not in wb.sheetnames:
                        needs_reset = True
                    else:
                        current_prod = [cell.value for cell in wb[PROD_SHEET][1]]
                        if current_prod != PROD_COLUMNS:
                            needs_reset = True
                    
                    # Check EVAL sheet
                    if EVAL_SHEET not in wb.sheetnames:
                        needs_reset = True
                    else:
                        current_eval = [cell.value for cell in wb[EVAL_SHEET][1]]
                        if current_eval != EVAL_COLUMNS:
                            needs_reset = True
                            
                except Exception:
                    needs_reset = True

            if needs_reset:
                if self.log_file.exists():
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    backup = self.log_file.with_name(f"response_log_STALE_{timestamp}.xlsx")
                    print(f"⚠ Self-healing: Excel columns outdated. Archiving to {backup.name}")
                    try:
                        shutil.move(str(self.log_file), str(backup))
                    except Exception as e:
                        print(f"  Failed to archive: {e}")
                
                self._make_fresh_wb().save(self.log_file)
                print(f"✓ Self-healing: Fresh log created with unified schema at {self.log_file}")

    def log_response(self, user_query, bot_response, time_taken,
                      session_id="N/A", source="N/A", context=None):
        """Append a row to the Production sheet with context pass-through."""
        with self.lock:
            try:
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                wb = load_workbook(self.log_file)
                
                if PROD_SHEET not in wb.sheetnames:
                    ws = wb.create_sheet(PROD_SHEET, 0)
                    ws.append(PROD_COLUMNS)
                else:
                    ws = wb[PROD_SHEET]

                # Clean context data
                if isinstance(context, list):
                    context_str = " | ".join([str(c) for c in context if c])
                else:
                    context_str = str(context) if context else "None"

                ws.append([
                    timestamp, user_query, bot_response,
                    f"{time_taken:.4f}", session_id, source, context_str
                ])
                wb.save(self.log_file)
                print(f"✓ Logged to Production: {source} query")
            except PermissionError:
                print(f"⚠ ERROR: Could not write to Excel. Is it open?")
            except Exception as e:
                print(f"Error logging: {e}")
