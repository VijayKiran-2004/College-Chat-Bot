import os
import sys
from pathlib import Path

# Add project root to sys.path to resolve 'app' module
project_root = str(Path(__file__).resolve().parent.parent)
if project_root not in sys.path:
    sys.path.append(project_root)

import requests
import pandas as pd
import time
import json
import re
from sentence_transformers import SentenceTransformer, util
from app.services.logger_service import ResponseLogger, EVAL_COLUMNS, EVAL_SHEET

API_URL    = "http://127.0.0.1:8000/query"
OLLAMA_URL = "http://127.0.0.1:11434/api/generate"
EVAL_FILE = "logs/response_log.xlsx"

# Initialize Similarity Model (for BERTScore-style eval)
print("Loading similarity model (all-MiniLM-L6-v2) for evaluation...")
sim_model = SentenceTransformer('all-MiniLM-L6-v2')

# Use centralized schema
EVAL_COLS = EVAL_COLUMNS

# --------------------------------------------------------------
# ALL YOUR PROMPTS (Loaded from JSON for scalability)
# --------------------------------------------------------------
PROMPTS_FILE = os.path.join(os.path.dirname(__file__), "all_prompts.json")

if os.path.exists(PROMPTS_FILE):
    with open(PROMPTS_FILE, 'r') as f:
        test_prompts = json.load(f)
    print(f"✓ Loaded {len(test_prompts)} prompts from {PROMPTS_FILE}")
else:
    print(f"⚠ {PROMPTS_FILE} not found! Using a minimal fallback list.")
    test_prompts = ["Who is the principal?", "What is the attendance criteria?", "List all CSE students."]

# LLM Judge function removed per user request. 
# Qualitative metrics are disabled in favor of objective statistical checks.

# --------------------------------------------------------------
# BERTSCORE: Token-level semantic overlap (answer vs context)
# --------------------------------------------------------------
def bertscore_f1(answer, context):
    """Compute a BERTScore-style F1 using sentence-transformers. No extra models needed."""
    if not answer or not context:
        return 0.5  # Neutral fallback for SQL/KB with no context
    
    ans_sents = [s.strip() for s in re.split(r'[.!?]', answer) if len(s.strip()) > 10]
    ctx_sents = [s.strip() for s in re.split(r'[.!?]', context) if len(s.strip()) > 10]

    if not ans_sents or not ctx_sents:
        return float(util.pytorch_cos_sim(
            sim_model.encode(answer, convert_to_tensor=True),
            sim_model.encode(context, convert_to_tensor=True)
        ).item())

    ans_embs = sim_model.encode(ans_sents, convert_to_tensor=True)
    ctx_embs = sim_model.encode(ctx_sents, convert_to_tensor=True)
    sim_matrix = util.pytorch_cos_sim(ans_embs, ctx_embs)

    precision = float(sim_matrix.max(dim=1).values.mean())
    recall    = float(sim_matrix.max(dim=0).values.mean())

    if precision + recall == 0: return 0.0
    return round(2 * precision * recall / (precision + recall), 4)

# --------------------------------------------------------------
# LINK VALIDITY
# --------------------------------------------------------------
def check_link_validity(text):
    links = re.findall(r'\[.*?\]\((https?://.*?)\)', text)
    if not links: return "N/A", 1.0
    valid_count = 0
    for link in links:
        try:
            r = requests.head(link, timeout=5, allow_redirects=True)
            if r.status_code < 400: valid_count += 1
            elif r.status_code in (403, 405, 406):
                r2 = requests.get(link, timeout=5, allow_redirects=True, stream=True)
                r2.close()
                if r2.status_code < 400: valid_count += 1
        except: continue
    return f"{valid_count}/{len(links)} Valid", round(valid_count / len(links), 4)

# --------------------------------------------------------------
# SOURCE APPROPRIATENESS
# --------------------------------------------------------------
def source_appropriateness(query, source):
    """Rule-based check: did the right system handle this query?"""
    q = query.lower()
    is_sql = any(kw in q for kw in ["how many", "list all", "who placed", "roll no", "cgpa", "student"])
    if is_sql and source == "SQL Database": return 1.0
    if not is_sql and source in ["Knowledge Base", "RAG"]: return 1.0
    return 0.5 if source != "Unknown" else 0.0

# --------------------------------------------------------------
# WAIT FOR BACKEND
# --------------------------------------------------------------
print("Checking backend status...")
while True:
    try:
        if requests.get("http://127.0.0.1:8000/health", timeout=3).status_code == 200:
            print("✓ Backend is online.\n")
            break
    except:
        print("⏳ Waiting for backend...")
        time.sleep(1)

# --------------------------------------------------------------
# EXCEL HELPERS (Incremental Save)
# --------------------------------------------------------------
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font

def _open_eval_sheet():
    if os.path.exists(EVAL_FILE): wb = load_workbook(EVAL_FILE)
    else: wb = Workbook()
    if EVAL_SHEET in wb.sheetnames: ws = wb[EVAL_SHEET]
    else:
        ws = wb.create_sheet(EVAL_SHEET)
        ws.append(EVAL_COLS)
        for cell in ws[1]: cell.font = Font(bold=True)
    return wb, ws

def _append_and_save(row_list: list):
    try:
        wb, ws = _open_eval_sheet()
        ws.append(row_list)
        wb.save(EVAL_FILE)
        print(f"  ✓ Saved to Evaluation sheet ({ws.max_row - 1} rows so far)")
    except Exception as e:
        print(f"  ⚠ Save error: {e}")

# --------------------------------------------------------------
# RUN TESTS
# --------------------------------------------------------------
results = []
for prompt in test_prompts:
    print(f"\n→ Testing: {prompt}")
    start = time.time()
    
    answer = ""
    source = "Unknown"
    confidence = "N/A"
    server_time = "N/A"
    context_list = []

    try:
        resp = requests.post(API_URL, json={"message": prompt}, timeout=60, stream=True)
        latency = round(time.time() - start, 3)

        if resp.status_code == 200:
            ctype = resp.headers.get("Content-Type", "")
            if "text/event-stream" in ctype:
                for line in resp.iter_lines():
                    if line:
                        line_str = line.decode('utf-8')
                        if line_str.startswith("data: "):
                            try:
                                d = json.loads(line_str[6:])
                                if d.get("type") == "metadata":
                                    confidence = d.get("confidence", "N/A")
                                    context_list = d.get("context", [])
                                    source = d.get("source", "RAG")
                                if "chunk" in d: answer += d["chunk"]
                                if d.get("done"):
                                    server_time = d.get("time_taken", "N/A")
                                    break
                            except: continue
            else:
                d = resp.json()
                answer = d.get("answer", "")
                source = d.get("source", "Unknown")
                confidence = d.get("confidence", "N/A")
                server_time = d.get("time_taken", "N/A")
                context_list = d.get("context", [])
        else: answer = f"ERROR ({resp.status_code})"
    except Exception as e:
        latency = round(time.time() - start, 3)
        answer = f"CONNECTION ERROR: {str(e)}"

    full_context = " | ".join(context_list) if context_list else ""
    
    # Metrics (BERT 50% | Source 25% | Links 25%)
    bert_f1 = bertscore_f1(answer, full_context)
    link_label, link_score = check_link_validity(answer)
    src_score = source_appropriateness(prompt, source)
    
    accuracy_pct = round((bert_f1 * 0.50 + src_score * 0.25 + link_score * 0.25) * 100, 2)
    bert_pct = round(bert_f1 * 100, 1)

    print(f"✓ Source: {source} | Latency: {latency}s | Server Time: {server_time}s")
    print(f"  [BERTScore] F1: {bert_pct}% | [Links] {link_label}")
    print(f"  [Final Accuracy] {accuracy_pct}%")

    row_data = [
        time.strftime("%Y-%m-%d %H:%M:%S"), prompt, answer, source,
        confidence, latency, server_time, bert_pct, link_label, accuracy_pct, full_context
    ]
    results.append(row_data)
    _append_and_save(row_data)
    time.sleep(0.5)

# --------------------------------------------------------------
# FINAL SUMMARY
# --------------------------------------------------------------
if results:
    df = pd.DataFrame(results, columns=EVAL_COLS)
    print(f"\n📊 Summary — Avg Accuracy: {df['Accuracy %'].mean():.1f}% | Avg Latency: {df['Latency (s)'].mean():.2f}s")
else:
    print("\n⚠ No results.")
